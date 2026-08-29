import hashlib
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
import unittest
import zipfile
from xml.etree import ElementTree as ET
from email import policy
from email.parser import BytesParser
from pathlib import Path, PurePosixPath


ROOT = Path(__file__).resolve().parents[1]
CATALOG_PATH = ROOT / "library/catalog.json"
TOKEN = re.compile(rb"\{\{([a-zA-Z0-9_]+)\}\}")
FORBIDDEN_NATIVE_RESIDUE = (
    b"syn studios",
    b"syn-studios",
    b"template",
    b"build-time",
    b"double-brace",
    b"instantiation gate",
    b"generator residue",
    b"template.invalid",
    b"2013-12-23",
)


def load_json(relative):
    return json.loads((ROOT / relative).read_text(encoding="utf-8"))


def sha256(path):
    return hashlib.sha256(path.read_bytes()).hexdigest()


def model_visible_payload(path):
    if path.suffix.lower() == ".xlsx":
        with zipfile.ZipFile(path) as archive:
            return b" ".join(
                archive.read(name)
                for name in archive.namelist()
                if name.startswith("xl/worksheets/")
                or name in {"xl/sharedStrings.xml", "xl/workbook.xml", "docProps/core.xml", "docProps/app.xml"}
            )
    if path.suffix.lower() == ".docx":
        with zipfile.ZipFile(path) as archive:
            return b" ".join(
                archive.read(name)
                for name in archive.namelist()
                if name == "word/document.xml"
                or name.startswith("word/header")
                or name.startswith("word/footer")
                or name.startswith("word/comments")
                or name in {"docProps/core.xml", "docProps/app.xml"}
            )
    if path.suffix.lower() == ".eml":
        raw = path.read_bytes()
        message = BytesParser(policy=policy.default).parsebytes(raw)
        chunks = [raw, *[str(value).encode("utf-8") for value in message.values()]]
        for part in message.walk():
            decoded = part.get_payload(decode=True)
            if decoded:
                chunks.append(decoded)
            if part.get_filename():
                chunks.append(part.get_filename().encode("utf-8"))
        return b" ".join(chunks)
    return path.read_bytes()


def tokens_in_asset(path):
    return {match.decode("ascii") for match in TOKEN.findall(model_visible_payload(path))}


def pdf_page_count(path):
    payload = path.read_bytes()
    if not payload.startswith(b"%PDF-") or b"%%EOF" not in payload[-1024:]:
        return None
    return len(re.findall(rb"/Type\s*/Page\b", payload))


def render_binding_valid(descriptor, manifest_override=None):
    contract = descriptor["render_contract"]
    if not contract["required"]:
        return contract["evidence_manifest"] is None and contract["expected_page_count"] is None
    manifest = manifest_override if manifest_override is not None else load_json(contract["evidence_manifest"])
    record = manifest["templates"].get(descriptor["template_id"])
    if not isinstance(record, dict) or len(descriptor["native_assets"]) != 1:
        return False
    asset = descriptor["native_assets"][0]
    if record.get("asset_path") != asset["path"] or record.get("asset_sha256") != asset["sha256"]:
        return False
    if record.get("page_count") != contract["expected_page_count"]:
        return False
    if record.get("sheet_names", []) != contract["expected_sheet_names"]:
        return False
    outputs = record.get("rendered_outputs")
    if not isinstance(outputs, list) or len(outputs) != record["page_count"] + 1:
        return False
    expected_paths = [contract["expected_pdf_path"]] + [
        contract["expected_page_image_pattern"].format(page=page)
        for page in range(1, record["page_count"] + 1)
    ]
    observed_paths = [item.get("path") for item in outputs]
    observed_hashes = [item.get("sha256") for item in outputs]
    if observed_paths != expected_paths:
        return False
    if len(set(observed_paths)) != len(observed_paths) or len(set(observed_hashes)) != len(observed_hashes):
        return False
    if PurePosixPath(observed_paths[0]).suffix.lower() != ".pdf":
        return False
    if any(PurePosixPath(path).suffix.lower() != ".png" for path in observed_paths[1:]):
        return False
    pdf_path = ROOT / observed_paths[0]
    if pdf_page_count(pdf_path) != record["page_count"]:
        return False
    return all((ROOT / item["path"]).is_file() and sha256(ROOT / item["path"]) == item["sha256"] for item in outputs)


def recalculate_workbook(mutator):
    try:
        import openpyxl
    except ImportError as error:
        raise unittest.SkipTest("openpyxl is unavailable for native workbook sabotage") from error
    executable = os.environ.get("SYN_STUDIOS_LIBREOFFICE") or shutil.which("soffice") or shutil.which("libreoffice")
    if not executable:
        standard = Path(r"C:\Program Files\LibreOffice\program\soffice.com")
        executable = str(standard) if standard.is_file() else None
    if not executable:
        raise unittest.SkipTest("LibreOffice is unavailable for native workbook recalculation")
    source = ROOT / "library/templates/TMPL-0001/1.0.0/internal-close-reconciliation.xlsx"
    with tempfile.TemporaryDirectory() as directory:
        root = Path(directory)
        input_dir, output_dir = root / "input", root / "output"
        input_dir.mkdir()
        output_dir.mkdir()
        working = input_dir / "sabotage.xlsx"
        shutil.copy2(source, working)
        workbook = openpyxl.load_workbook(working)
        mutator(workbook)
        workbook.save(working)
        result = subprocess.run(
            [executable, "--headless", f"-env:UserInstallation={(root / 'profile').as_uri()}", "--convert-to", "xlsx", "--outdir", str(output_dir), str(working)],
            check=False,
            capture_output=True,
            text=True,
            timeout=60,
        )
        if result.returncode:
            raise AssertionError(f"LibreOffice recalculation failed: {result.stdout}\n{result.stderr}")
        calculated = openpyxl.load_workbook(output_dir / working.name, data_only=True)
        checks = calculated["Checks"]
        return {cell: checks[cell].value for cell in ("B5", "B6", "B7", "B8", "B9", "B10", "B11", "B12", "B13", "B16", "C13")}


def populate_identity(workbook):
    from datetime import date

    control = workbook["Close_Control"]
    control["B5"], control["B6"] = "Example Organization", date(2026, 8, 31)
    control["B7"], control["B8"] = "Senior Accountant", "Controller"


def populate_minimal_valid_workbook(workbook):
    from datetime import date

    populate_identity(workbook)
    for column, value in enumerate(("ROW-1", "ENT-1", "1000", date(2026, 8, 31), "Balanced source row", 100, 100, "POSTED"), 1):
        workbook["Source_Data"].cell(5, column).value = value
    for column, value in enumerate(("1000", "Cash", "Assets", "Senior Accountant", "Yes"), 1):
        workbook["Account_Map"].cell(5, column).value = value
    for column, value in {1: "1000", 2: "Cash", 4: 0, 6: 0, 8: "No variance"}.items():
        workbook["Reconciliation"].cell(5, column).value = value


def table_bindings(path):
    bindings = {}
    with zipfile.ZipFile(path) as archive:
        for name in archive.namelist():
            if not name.startswith("xl/tables/") or not name.endswith(".xml"):
                continue
            root = ET.fromstring(archive.read(name))
            bindings[root.attrib["name"]] = root.attrib["ref"]
    return bindings


def population_binding_valid(descriptor, path):
    observed = table_bindings(path)
    for item in descriptor["population_contract"]["tables"]:
        if observed.get(item["name"]) != item["range"]:
            return False
        first, last = item["range"].split(":")
        first_row = int(re.search(r"[0-9]+$", first).group())
        last_row = int(re.search(r"[0-9]+$", last).group())
        if last_row - first_row != item["maximum_rows"]:
            return False
    return len(observed) == len(descriptor["population_contract"]["tables"])


class TemplateAssetTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.catalog = load_json("library/catalog.json")

    def test_catalog_has_three_versioned_discoverable_templates(self):
        entries = self.catalog["templates"]
        self.assertEqual(len(entries), 3)
        self.assertEqual({e["artifact_type"] for e in entries}, {"xlsx", "docx", "eml"})
        self.assertEqual({e["template_id"] for e in entries}, {"TMPL-0001", "TMPL-0002", "TMPL-0003"})
        required_facets = {"artifact_type", "authority", "lifecycle", "capabilities"}
        self.assertTrue(required_facets.issubset(self.catalog["discovery_fields"]))
        for entry in entries:
            self.assertRegex(entry["version"], r"^[0-9]+\.[0-9]+\.[0-9]+$")
            self.assertIn(entry["release_status"], {"candidate", "released"})
            self.assertTrue(entry["capabilities"])
            self.assertEqual(set(entry["supported_consumers"]), {"anna", "holodeck-file-generation", "human-artifact-realism"})

    def test_descriptors_and_assets_resolve_without_path_escape(self):
        for entry in self.catalog["templates"]:
            paths = [entry["descriptor"], *entry["native_assets"]]
            for relative in paths:
                pure = PurePosixPath(relative)
                self.assertFalse(pure.is_absolute(), relative)
                self.assertNotIn("..", pure.parts, relative)
                self.assertTrue((ROOT / relative).is_file(), relative)
            descriptor = load_json(entry["descriptor"])
            self.assertEqual(descriptor["template_id"], entry["template_id"])
            self.assertEqual(descriptor["version"], entry["version"])
            self.assertEqual(descriptor["blueprint_id"], entry["blueprint_id"])
            self.assertEqual(descriptor["artifact_type"], entry["artifact_type"])
            self.assertEqual(descriptor["authority"], entry["authority"])
            self.assertEqual(descriptor["lifecycle"], entry["lifecycle"])
            self.assertEqual(descriptor["release_status"], entry["release_status"])
            self.assertEqual(descriptor["capabilities"], entry["capabilities"])
            self.assertEqual(descriptor["supported_consumers"], entry["supported_consumers"])
            self.assertEqual([asset["path"] for asset in descriptor["native_assets"]], entry["native_assets"])

    def test_descriptors_reference_blueprints_without_duplicating_lineage(self):
        for entry in self.catalog["templates"]:
            descriptor = load_json(entry["descriptor"])
            self.assertEqual(descriptor["blueprint_id"], entry["blueprint_id"])
            self.assertNotIn("lineage", descriptor)

    def test_catalog_authority_sabotage_breaks_descriptor_binding(self):
        entry = dict(self.catalog["templates"][0])
        entry["authority"] = "authoritative"
        descriptor = load_json(entry["descriptor"])
        self.assertNotEqual(entry["authority"], descriptor["authority"])

    def test_native_asset_hashes_match_descriptors(self):
        for entry in self.catalog["templates"]:
            descriptor = load_json(entry["descriptor"])
            self.assertEqual(len(descriptor["native_assets"]), len(entry["native_assets"]))
            for asset in descriptor["native_assets"]:
                path = ROOT / asset["path"]
                self.assertEqual(sha256(path), asset["sha256"], asset["path"])

    def test_hash_gate_detects_single_byte_sabotage(self):
        descriptor = load_json("library/templates/TMPL-0003/1.0.0/template.json")
        asset = descriptor["native_assets"][0]
        original = ROOT / asset["path"]
        with tempfile.TemporaryDirectory() as directory:
            sabotaged = Path(directory) / original.name
            sabotaged.write_bytes(original.read_bytes() + b"\x00")
            self.assertNotEqual(sha256(sabotaged), asset["sha256"])

    def test_render_manifest_binds_current_asset_and_every_output_hash(self):
        for entry in self.catalog["templates"]:
            descriptor = load_json(entry["descriptor"])
            self.assertTrue(render_binding_valid(descriptor), entry["template_id"])

    def test_render_evidence_rejects_asset_or_descriptor_drift(self):
        descriptor = load_json("library/templates/TMPL-0002/1.0.0/template.json")
        descriptor["native_assets"][0]["sha256"] = "0" * 64
        self.assertFalse(render_binding_valid(descriptor))

    def test_render_evidence_rejects_duplicate_page_substitution(self):
        descriptor = load_json("library/templates/TMPL-0002/1.0.0/template.json")
        manifest = load_json(descriptor["render_contract"]["evidence_manifest"])
        record = manifest["templates"][descriptor["template_id"]]
        record["rendered_outputs"][2] = dict(record["rendered_outputs"][1])
        self.assertFalse(render_binding_valid(descriptor, manifest))

    def test_every_embedded_build_token_is_declared(self):
        for entry in self.catalog["templates"]:
            descriptor = load_json(entry["descriptor"])
            declared = set(descriptor["slots"])
            observed = set()
            for asset in descriptor["native_assets"]:
                observed.update(tokens_in_asset(ROOT / asset["path"]))
            self.assertEqual(observed, declared, entry["template_id"])

    def test_native_assets_contain_no_library_or_build_residue(self):
        for entry in self.catalog["templates"]:
            for relative in entry["native_assets"]:
                payload = model_visible_payload(ROOT / relative).lower()
                for residue in FORBIDDEN_NATIVE_RESIDUE:
                    self.assertNotIn(residue, payload, f"{relative}: {residue!r}")

    def test_required_vertical_slice_capabilities(self):
        required = {
            "TMPL-0001": {"native_xlsx", "formula_reconciliation", "exception_log", "control_checks", "renderable"},
            "TMPL-0002": {"native_docx", "issue_analysis", "risk_and_sensitivity", "open_items", "renderable"},
            "TMPL-0003": {"native_eml", "multipart_mime", "quoted_thread", "attachment_correction", "machine_parseable"},
        }
        for entry in self.catalog["templates"]:
            self.assertTrue(required[entry["template_id"]].issubset(entry["capabilities"]))

    def test_xlsx_has_expected_layers_and_formula_controls(self):
        path = ROOT / "library/templates/TMPL-0001/1.0.0/internal-close-reconciliation.xlsx"
        with zipfile.ZipFile(path) as archive:
            workbook_xml = archive.read("xl/workbook.xml")
            formulas = b" ".join(archive.read(name) for name in archive.namelist() if name.startswith("xl/worksheets/") and name.endswith(".xml"))
        for name in (b"Close_Control", b"Source_Data", b"Account_Map", b"Reconciliation", b"Proposed_Entries", b"Prior_Period", b"Exceptions", b"Checks"):
            self.assertIn(name, workbook_xml)
        self.assertEqual(workbook_xml.count(b"_xlnm.Print_Area"), 8)
        self.assertEqual(formulas.count(b"fitToWidth=\"1\""), 8)
        self.assertGreaterEqual(formulas.count(b"<pane"), 7)
        self.assertGreaterEqual(formulas.count(b"tablePart"), 6)
        self.assertIn(b"SUMIFS", formulas)
        self.assertIn(b"COUNTIF", formulas)
        with zipfile.ZipFile(path) as archive:
            all_xml = b" ".join(archive.read(name) for name in archive.namelist() if name.endswith(".xml"))
        self.assertIn(b"{{organization_name}}", all_xml)

    def test_xlsx_population_contract_binds_table_capacity(self):
        descriptor = load_json("library/templates/TMPL-0001/1.0.0/template.json")
        path = ROOT / descriptor["native_assets"][0]["path"]
        self.assertTrue(population_binding_valid(descriptor, path))
        self.assertEqual(descriptor["population_contract"]["capacity_change_policy"], "reject_and_rebuild_with_fresh_formula_and_render_evidence")
        for item in descriptor["population_contract"]["tables"]:
            self.assertGreaterEqual(item["maximum_rows"], item["minimum_rows"])
            self.assertTrue(item["columns"])

    def test_xlsx_population_contract_supports_variable_native_and_csv_scale(self):
        descriptor = load_json("library/templates/TMPL-0001/1.0.0/template.json")
        contract = descriptor["population_contract"]
        carriers = {carrier["kind"]: carrier for carrier in contract["source_carriers"]}
        self.assertEqual(set(carriers), {"native_table", "csv_import"})
        self.assertTrue(carriers["native_table"]["variable_row_counts"])
        self.assertEqual(carriers["csv_import"]["target_table"], "SourceDataTable")
        self.assertEqual(set(carriers["csv_import"]["required_target_columns"]), set(contract["tables"][0]["columns"]))
        expansion = contract["expansion_contract"]
        self.assertTrue(expansion["reference_capacity_only"])
        self.assertEqual((expansion["minimum_population_rows"], expansion["reference_population_rows"], expansion["expanded_proof_rows"]), (1, 25, 30))
        for key in ("builder", "rebuild_pipeline", "evidence_verifier", "deterministic_test_carrier"):
            self.assertTrue((ROOT / expansion[key]).is_file(), key)
        self.assertIn("formula", expansion["formula_propagation"])
        self.assertIn("print", expansion["print_propagation"])
        self.assertEqual(set(expansion["proof_required"]), {"typed row validation", "formula recalculation", "out-of-capacity scan", "all-sheet render", "fresh native and render hashes"})
        self.assertTrue({"variable_row_counts", "csv_import_compatible", "rebuildable_capacity"}.issubset(descriptor["capabilities"]))

    def test_xlsx_population_expansion_without_rebuild_is_rejected(self):
        descriptor = load_json("library/templates/TMPL-0001/1.0.0/template.json")
        original = ROOT / descriptor["native_assets"][0]["path"]
        with tempfile.TemporaryDirectory() as directory:
            expanded = Path(directory) / original.name
            with zipfile.ZipFile(original) as source, zipfile.ZipFile(expanded, "w") as target:
                for info in source.infolist():
                    payload = source.read(info.filename)
                    if info.filename.startswith("xl/tables/") and info.filename.endswith(".xml"):
                        table = ET.fromstring(payload)
                        if table.attrib.get("name") == "SourceDataTable":
                            table.attrib["ref"] = "A4:H30"
                            payload = ET.tostring(table, encoding="utf-8", xml_declaration=True)
                    target.writestr(info, payload)
            self.assertFalse(population_binding_valid(descriptor, expanded))

    def test_xlsx_empty_rows_are_not_false_passes_and_readiness_has_no_tokens(self):
        path = ROOT / "library/templates/TMPL-0001/1.0.0/internal-close-reconciliation.xlsx"
        with zipfile.ZipFile(path) as archive:
            worksheets = b" ".join(archive.read(name) for name in archive.namelist() if name.startswith("xl/worksheets/") and name.endswith(".xml"))
            checks = ET.fromstring(archive.read("xl/worksheets/sheet8.xml"))
        ns = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        formula_text = []
        with zipfile.ZipFile(path) as archive:
            for name in archive.namelist():
                if name.startswith("xl/worksheets/") and name.endswith(".xml"):
                    formula_text.extend(element.text or "" for element in ET.fromstring(archive.read(name)).findall(".//m:f", ns))
        formulas = " ".join(formula_text).encode("utf-8")
        self.assertNotIn(b"{{", formulas)
        self.assertIn(b'IF(A5="",""', formulas)
        self.assertIn(b'"NOT READY"', formulas)
        self.assertIn(b'COUNTIF(B5:B13,"FAIL")>0', formulas)
        self.assertNotIn(b'=0,"PASS",IF(ABS(C9)', formulas)
        cached = []
        for ref in ["B5", "B6", "B7", "B8", "B9", "B10", "B11", "B12", "B13", "B16"]:
            value = checks.find(f".//m:c[@r='{ref}']/m:v", ns)
            cached.append(value.text if value is not None else "")
        self.assertNotIn("PASS", cached[:8])
        self.assertEqual(cached[-2], "PASS")
        self.assertEqual(cached[-1], "NOT READY")

    def test_xlsx_shallow_stub_recalculates_to_fail(self):
        def shallow_stub(workbook):
            populate_identity(workbook)
            workbook["Source_Data"]["A5"] = "ROW-1"
            workbook["Account_Map"]["A5"] = "1000"
            reconciliation = workbook["Reconciliation"]
            reconciliation["A5"], reconciliation["D5"], reconciliation["F5"] = "1000", 0, 0

        checks = recalculate_workbook(shallow_stub)
        self.assertEqual(checks["B5"], "FAIL")
        self.assertEqual(checks["B8"], "FAIL")
        self.assertEqual(checks["B9"], "FAIL")
        self.assertEqual(checks["B16"], "FAIL")

    def test_xlsx_populated_minimal_recalculates_to_pass(self):
        checks = recalculate_workbook(populate_minimal_valid_workbook)
        self.assertEqual(checks["B10"], "NOT APPLICABLE")
        self.assertEqual(checks["B16"], "PASS")

    def test_xlsx_unmatched_source_account_recalculates_to_fail(self):
        def unmatched_mapping(workbook):
            populate_minimal_valid_workbook(workbook)
            workbook["Account_Map"]["A5"] = "DIFFERENT"

        checks = recalculate_workbook(unmatched_mapping)
        self.assertEqual(checks["B8"], "FAIL")
        self.assertEqual(checks["B16"], "FAIL")

    def test_xlsx_malformed_exception_recalculates_to_fail(self):
        def malformed_exception(workbook):
            populate_minimal_valid_workbook(workbook)
            workbook["Exceptions"]["A5"] = "EX-1"
            workbook["Exceptions"]["C5"] = "Unresolved"
            workbook["Exceptions"]["G5"] = "Pending"

        checks = recalculate_workbook(malformed_exception)
        self.assertEqual(checks["B11"], "FAIL")
        self.assertEqual(checks["B16"], "FAIL")

    def test_xlsx_partial_proposed_entry_recalculates_to_fail(self):
        def partial_entry(workbook):
            populate_minimal_valid_workbook(workbook)
            workbook["Proposed_Entries"]["D5"] = 500

        checks = recalculate_workbook(partial_entry)
        self.assertEqual(checks["B10"], "FAIL")
        self.assertEqual(checks["B16"], "FAIL")

    def test_xlsx_csv_population_rebuilds_and_binds_expanded_render_evidence(self):
        node = os.environ.get("SYN_STUDIOS_NODE") or shutil.which("node")
        if not node or not os.environ.get("SYN_STUDIOS_NODE_MODULES"):
            raise unittest.SkipTest("activated document generation stack is unavailable")
        soffice = os.environ.get("SYN_STUDIOS_SOFFICE")
        poppler = os.environ.get("SYN_STUDIOS_POPPLER_BIN")
        if not soffice or not poppler:
            raise unittest.SkipTest("activated render stack is unavailable")

        pipeline = ROOT / "evidence/reports/template-assets/builders/build_close_population.py"
        fixture = ROOT / "tests/fixtures/populations/source-expanded-30.csv"
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            rebuilt = root / "expanded.xlsx"
            evidence_dir = root / "evidence"
            subprocess.run(
                [sys.executable, str(pipeline), "--source-csv", str(fixture), "--output", str(rebuilt), "--evidence-dir", str(evidence_dir)],
                check=True,
                timeout=180,
            )
            evidence_path = evidence_dir / "population-evidence.json"
            evidence = json.loads(evidence_path.read_text(encoding="utf-8"))
            self.assertEqual(evidence["source_table_ref"], "A4:H34")
            self.assertEqual(evidence["source_print_area"], "'Source_Data'!$A$1:$H$34")
            self.assertEqual(evidence["formula_boundary"], 34)
            self.assertEqual(evidence["control_results"], {"B5": "PASS", "B6": "PASS", "B7": "PASS", "B8": "NOT READY", "B13": "PASS", "B16": "NOT READY"})
            self.assertEqual(evidence["verdict"], "PASS")
            self.assertGreaterEqual(len(evidence["rendered_outputs"]), 2)
            self.assertTrue(all(re.fullmatch(r"[0-9a-f]{64}", item["sha256"]) for item in evidence["rendered_outputs"]))

    def test_xlsx_csv_population_rejects_incomplete_typed_row(self):
        node = os.environ.get("SYN_STUDIOS_NODE") or shutil.which("node")
        if not node or not os.environ.get("SYN_STUDIOS_NODE_MODULES"):
            raise unittest.SkipTest("activated document generation stack is unavailable")
        builder = ROOT / "evidence/reports/template-assets/builders/build_close_template.mjs"
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            incomplete = root / "incomplete.csv"
            incomplete.write_text(
                "Source Row ID,Entity Code,Account Code,Transaction Date,Description,Debit,Credit,Status Code\n"
                "ROW-1,ENTITY-A,1000,2026-08-31,Incomplete row,100.00,,POSTED\n",
                encoding="utf-8",
            )
            completed = subprocess.run(
                [node, str(builder), str(root / "rejected.xlsx"), "--source-csv", str(incomplete)],
                check=False,
                capture_output=True,
                text=True,
                timeout=60,
            )
            self.assertNotEqual(completed.returncode, 0)
            self.assertIn("must contain all eight fields", completed.stderr)

    def test_xlsx_out_of_capacity_cell_recalculates_to_fail(self):
        def out_of_capacity(workbook):
            populate_minimal_valid_workbook(workbook)
            workbook["Source_Data"]["F30"] = 999

        checks = recalculate_workbook(out_of_capacity)
        self.assertEqual(checks["C13"], 1)
        self.assertEqual(checks["B13"], "FAIL")
        self.assertEqual(checks["B16"], "FAIL")

    def test_docx_has_five_renderable_sections_and_no_approval_claim(self):
        path = ROOT / "library/templates/TMPL-0002/1.0.0/internal-controller-memo.docx"
        with zipfile.ZipFile(path) as archive:
            document = archive.read("word/document.xml")
            core = archive.read("docProps/core.xml")
            app = archive.read("docProps/app.xml")
        self.assertGreaterEqual(document.count(b'w:type="page"'), 4)
        self.assertIn(b"NOT APPROVED", document)
        self.assertNotIn(b"dcterms:created", core)
        self.assertNotIn(b"dc:creator", core)
        self.assertNotIn(b"Application", app)

    def test_eml_is_multipart_with_two_openable_attachments(self):
        path = ROOT / "library/templates/TMPL-0003/1.0.0/operational-correction-thread.eml"
        message = BytesParser(policy=policy.default).parsebytes(path.read_bytes())
        self.assertTrue(message.is_multipart())
        self.assertFalse(any(name.lower().startswith("x-syn") for name in message.keys()))
        self.assertNotIn("syn", message.get_boundary().lower())
        self.assertNotIn("template", message.get_boundary().lower())
        body = message.get_body(preferencelist=("plain",)).get_content()
        self.assertTrue(body.startswith("Team,"))
        self.assertNotIn("{{operations_manager_first_name}}", body)
        self.assertEqual(body.count("-----Original Message-----") + 1, 6)
        attachments = list(message.iter_attachments())
        self.assertEqual(len(attachments), 2)
        self.assertEqual({part.get_content_type() for part in attachments}, {"text/csv", "text/plain"})
        for part in attachments:
            self.assertTrue(part.get_payload(decode=True))

    def test_mixed_manifest_references_exact_catalog_versions(self):
        manifest = load_json("examples/manifests/mixed-controller-review-packet.json")
        catalog_pairs = {(item["template_id"], item["version"]) for item in self.catalog["templates"]}
        manifest_pairs = {(item["template_id"], item["version"]) for item in manifest["components"]}
        self.assertEqual(manifest_pairs, catalog_pairs)
        self.assertTrue(manifest["required_additional_components"])


if __name__ == "__main__":
    unittest.main()
